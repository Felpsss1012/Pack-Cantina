# app.py
from flask import Flask, jsonify, render_template, send_from_directory, request, abort
from flask_socketio import SocketIO, emit
from flask_cors import CORS
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import json, os, re
import threading
import pathlib
import glob
import difflib
import unicodedata
from datetime import datetime, date, timedelta

# lock pra evitar escrita concorrente no arquivo de pedidos / produtos / planilhas
file_lock = threading.Lock()
PROJECT_ROOT = pathlib.Path(__file__).parent.resolve()

# ---------------- Configurações ----------------
EXCEL_DIR_FILES = []

HTML_DIR = PROJECT_ROOT / "html"
STYLE_DIR = PROJECT_ROOT / "Style"
SCRIPT_DIR = PROJECT_ROOT / "Script"
DATA_DIR = PROJECT_ROOT / "Data"
RELATORIO_DIR = PROJECT_ROOT / "Relatorios"
PRODUCTS_FILE = os.path.join(DATA_DIR, "products.json")
PEDIDOS_FILE = os.path.join(DATA_DIR, "pedidos.xlsx")
ENTRADAS_FILE = os.path.join(DATA_DIR, "entradas.xlsx")
SAIDAS_FILE = os.path.join(DATA_DIR, "saidas.xlsx")

os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__, static_folder="static", template_folder="templates")
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="eventlet")

@app.route("/", methods=["GET"])
def serve_index():
    """Serve html/index.html"""
    idx = HTML_DIR / "index.html"
    if idx.exists():
        return send_from_directory(str(HTML_DIR), "index.html")
    else:
        abort(404, description="index.html não encontrado")

@app.route("/painel", methods=["GET"])
def serve_painel():
    """Serve html/painel.html"""
    p = HTML_DIR / "painel.html"
    if p.exists():
        return send_from_directory(str(HTML_DIR), "painel.html")
    else:
        abort(404, description="painel.html não encontrado")

@app.route("/files/<path:filename>", methods=["GET"])
def serve_any_file(filename):
    """
    Serve arquivos estáticos/suplentes de forma genérica.
    Uso exemplo: /files/Style/styles.css  -> entrega Style/styles.css
                 /files/Script/script.js   -> entrega Script/script.js
                 /files/Data/saidas.xlsx   -> entrega Data/saidas.xlsx
    """
    # proteger caminho para não subir acima do projeto
    requested = pathlib.Path(filename)
    if requested.is_absolute() or ".." in requested.parts:
        abort(400, description="Caminho inválido")

    # tenta localizar nas pastas conhecidas (ordem: html, Style, Script, Data, Relatorios)
    candidates = [
        (HTML_DIR, filename),
        (STYLE_DIR, filename),
        (SCRIPT_DIR, filename),
        (DATA_DIR, filename),
        (RELATORIO_DIR, filename),
    ]
    for base, fn in candidates:
        fpath = base / fn
        if fpath.exists() and fpath.is_file():
            return send_from_directory(str(base), fn)

    # se não achou, tenta dentro do projeto (para cobrir subpastas)
    f_try = PROJECT_ROOT / filename
    if f_try.exists() and f_try.is_file():
        return send_from_directory(str(PROJECT_ROOT), filename)

    abort(404, description=f"Arquivo não encontrado: {filename}")

# atalhos úteis
@app.route("/data/<path:filename>", methods=["GET"])
def serve_data_file(filename):
    if ".." in filename:
        abort(400)
    f = DATA_DIR / filename
    if f.exists() and f.is_file():
        return send_from_directory(str(DATA_DIR), filename)
    abort(404)

# ---------------- Socket.IO handlers e endpoints mínimos ----------------
# Cole após a criação de `socketio` e das funções utilitárias de arquivo (append_rows_to_excel, save_products, load_products)

@socketio.on('connect')
def sock_connect():
    sid = request.sid if hasattr(request, "sid") else None
    print(f"[SOCKET] client connected sid={sid}")
    emit('server_message', {"msg": "connected", "sid": sid})

@socketio.on('disconnect')
def sock_disconnect():
    try:
        sid = request.sid if hasattr(request, "sid") else None
        print(f"[SOCKET] client disconnected sid={sid}")
    except Exception:
        pass

@socketio.on('novo_pedido')
def sock_novo_pedido(payload):
    """
    Espera payload no formato:
    { buyerInfo: { name }, itens: [{ id?, name, qtd, price? }], total, pagamento, note }
    """
    try:
        print("[SOCKET] novo_pedido recebido:", payload)
        # tenta reutilizar process_order_payload se existir
        if 'process_order_payload' in globals() and callable(process_order_payload):
            pedido = process_order_payload(payload)
        else:
            # fallback minimal: grava em PEDIDOS_FILE e em SAIDAS_FILE e atualiza products.json
            pedido_id = int(datetime.now().timestamp() * 1000)
            hora = datetime.now().isoformat(sep=' ', timespec='seconds')
            pedido = {
                "id": pedido_id,
                "nome_cliente": (payload.get("buyerInfo") or {}).get("name") or payload.get("nome_cliente") or "",
                "itens": payload.get("itens") or payload.get("itens") or [],
                "total": payload.get("total"),
                "pagamento": payload.get("pagamento") or payload.get("payment") or payload.get("payment_method",""),
                "status": "novo",
                "hora_pedido": hora,
                "hora_conclusao": ""
            }
            # salva pedido (usa função existente garantir_arquivo_pedidos e salvar_pedido_excel se disponível)
            try:
                if 'salvar_pedido_excel' in globals() and callable(salvar_pedido_excel):
                    salvar_pedido_excel(pedido)
                else:
                    append_rows_to_excel(PEDIDOS_FILE, [pedido])
            except Exception as e:
                print("[WARN] falha ao salvar pedido:", e)

            # registra saidas por item e decrementa produto em products.json
            products = load_products()
            prod_map = {p['id']: p for p in products}
            saidas_rows = []
            pagamento = pedido.get("pagamento", "")
            for it in pedido['itens']:
                name = it.get('name') or it.get('nome')
                qty = int(it.get('qtd', it.get('qty', 1)) or 1)
                # tenta achar por id
                pid = it.get('id')
                prod = None
                if pid is not None:
                    prod = prod_map.get(pid)
                if not prod:
                    # procura por nome exato (case-insensitive)
                    for p in products:
                        if (p.get('name') or '').strip().lower() == (name or '').strip().lower():
                            prod = p
                            break
                prev_qty = int(prod.get('quantity', 0) or 0) if prod else None
                new_qty = prev_qty - qty if prev_qty is not None else None
                if prod and new_qty is not None:
                    prod['quantity'] = max(0, new_qty)
                # registra saida
                saidas_rows.append({
                    "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "product_id": prod.get('id') if prod else pid,
                    "product_name": prod.get('name') if prod else name,
                    "qty": qty,
                    "unit_price": prod.get('price') if prod else (it.get('price') or 0),
                    "total_price": round(qty * (prod.get('price') if prod else (it.get('price') or 0)), 4),
                    "prev_qty": prev_qty,
                    "new_qty": prod.get('quantity') if prod else None,
                    "payment_method": pagamento,
                })
            try:
                if saidas_rows:
                    append_rows_to_excel(SAIDAS_FILE, saidas_rows, columns_order=[
                        "date","product_id","product_name","qty","unit_price","total_price","prev_qty","new_qty","payment_method"
                    ])
                # salva products.json atualizado
                save_products(products)
            except Exception as e:
                print("[WARN] falha ao gravar saidas/products:", e)

        # emit para painel/clients que um pedido foi recebido
        socketio.emit('pedido_recebido', pedido, broadcast=True)
        # confirma para quem enviou
        emit('pedido_confirmado', {"id": pedido.get("id"), "hora": pedido.get("hora_pedido")})
    except Exception as ex:
        print("[ERROR] sock_novo_pedido:", ex)
        try:
            emit('pedido_error', {"error": str(ex)})
        except Exception:
            pass

# HTTP fallback: POST /api/orders  (permite enviar pedido via HTTP)
@app.route('/api/orders', methods=['POST'])
def api_orders_post():
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({"error":"payload invalido"}), 400
    try:
        # reutiliza a mesma lógica do socket: se process_order_payload existe, usa
        if 'process_order_payload' in globals() and callable(process_order_payload):
            pedido = process_order_payload(data)
        else:
            # invoca handler socket novo_pedido para reaproveitar logica
            sock_novo_pedido(data)
            pedido = {"ok": True}
        return jsonify({"ok": True, "order": pedido}), 201
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------- Utilitários ----------------
def load_products():
    """Carrega o catálogo de produtos do JSON, criando se não existir."""
    if os.path.exists(PRODUCTS_FILE):
        with open(PRODUCTS_FILE, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except Exception:
                # se arquivo corrompido, renomeia e recria
                try:
                    bak = PRODUCTS_FILE + f".corrompido-{int(datetime.now().timestamp())}.bak"
                    os.replace(PRODUCTS_FILE, bak)
                    print(f"[WARN] products.json corrompido renomeado para {bak}")
                except Exception:
                    pass
    # inicial padrão se não existe
    initial = [
        {"id":1,"name":"Água com Gás","price":2.5,"quantity":10,"category":"Bebidas","avg_cost":2.5},
        {"id":2,"name":"Água Mineral","price":2.0,"quantity":10,"category":"Bebidas","avg_cost":2.0},
        {"id":3,"name":"Amendoim Mendorato","price":1.75,"quantity":10,"category":"Salgados","avg_cost":1.75},
        # ... aqui mantive sua lista inicial reduzida por legibilidade; o código tem o resto no seu arquivo original
    ]
    # se quiser manter a lista completa, pode inicializar com seus 50 itens; para brevidade deixei alguns
    with open(PRODUCTS_FILE,"w",encoding="utf-8") as f:
        json.dump(initial,f,ensure_ascii=False,indent=2)
    return initial

def save_products(products):
    with file_lock:
        with open(PRODUCTS_FILE,"w",encoding="utf-8") as f:
            json.dump(products,f,ensure_ascii=False,indent=2)

def garantir_arquivo_pedidos():
    """Cria arquivo pedidos.xlsx se não existir, usando engine explicitamente."""
    if not os.path.exists(PEDIDOS_FILE):
        df = pd.DataFrame(columns=[
            "id","nome_cliente","itens","total","pagamento",
            "status","hora_pedido","hora_conclusao"
        ])
        try:
            df.to_excel(PEDIDOS_FILE, index=False, engine="openpyxl")
        except Exception as e:
            fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
            df.to_csv(fallback, index=False, encoding='utf-8-sig')
            print(f"[WARN] Não foi possível criar {PEDIDOS_FILE} ({e}). Criado fallback CSV: {fallback}")

def reparar_arquivo_pedidos():
    try:
        pd.read_excel(PEDIDOS_FILE, engine="openpyxl")
        print("[INFO] pedidos.xlsx válido.")
        return
    except Exception as e:
        print(f"[WARN] pedidos.xlsx inválido ou inacessível: {e}")

    csv_fallback = PEDIDOS_FILE.replace(".xlsx", ".csv")
    if os.path.exists(csv_fallback):
        try:
            print(f"[INFO] Convertendo {csv_fallback} -> {PEDIDOS_FILE} ...")
            df = pd.read_csv(csv_fallback, dtype=str)
            df.to_excel(PEDIDOS_FILE, index=False, engine="openpyxl")
            print("[INFO] Conversão concluída.")
            return
        except Exception as ce:
            print(f"[ERRO] Falha ao converter CSV para XLSX: {ce}")

    if os.path.exists(PEDIDOS_FILE):
        try:
            bak = PEDIDOS_FILE.replace(".xlsx", f".corrompido-{int(datetime.now().timestamp())}.bak.xlsx")
            os.replace(PEDIDOS_FILE, bak)
            print(f"[WARN] Renomeado {PEDIDOS_FILE} para {bak}")
        except Exception as re:
            print(f"[WARN] Não foi possível renomear {PEDIDOS_FILE}: {re}")

    df = pd.DataFrame(columns=[
        "id","nome_cliente","itens","total","pagamento",
        "status","hora_pedido","hora_conclusao"
    ])
    try:
        df.to_excel(PEDIDOS_FILE, index=False, engine="openpyxl")
        print("[INFO] Arquivo pedidos.xlsx criado novo.")
    except Exception as ce:
        print(f"[ERRO] Não conseguiu criar pedidos.xlsx: {ce}. Verifique permissões.")

def salvar_pedido_excel(pedido):
    garantir_arquivo_pedidos()
    with file_lock:
        try:
            df = pd.read_excel(PEDIDOS_FILE, engine="openpyxl")
        except Exception as e_read:
            csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
            print(f"[WARN] Falha ao ler {PEDIDOS_FILE} com openpyxl: {e_read}. Tentando fallback CSV {csv_fallback}")
            if os.path.exists(csv_fallback):
                try:
                    df = pd.read_csv(csv_fallback, dtype=str)
                except Exception as e_csv_read:
                    print(f"[ERRO] Também falhou ler CSV fallback: {e_csv_read}. Criando DataFrame vazio.")
                    df = pd.DataFrame(columns=["id","nome_cliente","itens","total","pagamento","status","hora_pedido","hora_conclusao"])
            else:
                df = pd.DataFrame(columns=["id","nome_cliente","itens","total","pagamento","status","hora_pedido","hora_conclusao"])

        df = pd.concat([df, pd.DataFrame([pedido])], ignore_index=True)
        try:
            df.to_excel(PEDIDOS_FILE, index=False, engine="openpyxl")
        except Exception as e_write:
            csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
            try:
                df.to_csv(csv_fallback, index=False, encoding='utf-8-sig')
                print(f"[WARN] Não foi possível gravar .xlsx ({e_write}). Gravado CSV fallback em: {csv_fallback}")
            except Exception as e_csv:
                print(f"[ERRO] Falha ao gravar fallback CSV: {e_csv}")
                raise

def normalize_text(s):
    if s is None:
        return ""
    s = str(s)
    s = s.strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = "".join([c for c in s if not unicodedata.combining(c)])
    s = re.sub(r'\s+', ' ', s)
    return s

def load_manual_mappings():
    mp_file = os.path.join(DATA_DIR, "manual_mappings.json")
    if os.path.exists(mp_file):
        try:
            with open(mp_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print("[WARN] erro ao ler manual_mappings.json:", e)
            return {}
    return {}

MANUAL_MAPPINGS = load_manual_mappings()

# ---------------- Relatório Semanal (Segunda-Sexta) ----------------
# Cria pasta de relatórios se necessário
os.makedirs(RELATORIO_DIR, exist_ok=True)

def _read_saidas():
    """Lê Data/saidas.xlsx (ou fallback CSV). Retorna DataFrame com colunas normalizadas."""
    if os.path.exists(SAIDAS_FILE):
        try:
            df = pd.read_excel(SAIDAS_FILE, engine="openpyxl", dtype=str)
        except Exception as e:
            print(f"[WARN] Falha ao ler {SAIDAS_FILE}: {e}. Tentando CSV fallback.")
            csv_fallback = SAIDAS_FILE.replace('.xlsx', '.csv')
            if os.path.exists(csv_fallback):
                df = pd.read_csv(csv_fallback, dtype=str)
            else:
                return pd.DataFrame()
    else:
        csv_fallback = SAIDAS_FILE.replace('.xlsx', '.csv')
        if os.path.exists(csv_fallback):
            df = pd.read_csv(csv_fallback, dtype=str)
        else:
            return pd.DataFrame()

    # sanitize column names
    df.columns = [str(c).strip() for c in df.columns]

    # normaliza coluna de data
    if 'date' in df.columns:
        df['date_dt'] = pd.to_datetime(df['date'], errors='coerce')
    elif 'data' in df.columns:
        df['date_dt'] = pd.to_datetime(df['data'], errors='coerce')
    else:
        possible = [c for c in df.columns if 'data' in c.lower() or 'date' in c.lower()]
        if possible:
            df['date_dt'] = pd.to_datetime(df[possible[0]], errors='coerce')
        else:
            df['date_dt'] = pd.NaT

    # normaliza quantidade (qty_num)
    if 'qty' in df.columns:
        df['qty_num'] = pd.to_numeric(df['qty'], errors='coerce').fillna(0).astype(int)
    elif 'qtd' in df.columns:
        df['qty_num'] = pd.to_numeric(df['qtd'], errors='coerce').fillna(0).astype(int)
    elif 'quantidade' in df.columns:
        df['qty_num'] = pd.to_numeric(df['quantidade'], errors='coerce').fillna(0).astype(int)
    else:
        found = None
        for c in df.columns:
            if 'qty' in c.lower() or 'qtd' in c.lower() or 'quant' in c.lower():
                found = c; break
        if found:
            df['qty_num'] = pd.to_numeric(df[found], errors='coerce').fillna(0).astype(int)
        else:
            df['qty_num'] = 0

    # normaliza product id / name
    if 'product_id' not in df.columns and 'produto_id' in df.columns:
        df = df.rename(columns={'produto_id': 'product_id'})
    if 'product_name' not in df.columns:
        if 'produto' in df.columns:
            df = df.rename(columns={'produto': 'product_name'})
        elif 'product' in df.columns:
            df = df.rename(columns={'product': 'product_name'})

    # garante colunas existirem
    if 'product_id' not in df.columns:
        df['product_id'] = None
    if 'product_name' not in df.columns:
        df['product_name'] = ''

    # payment method normalization (se existir mantém como string)
    if 'payment_method' not in df.columns and 'pagamento' in df.columns:
        df = df.rename(columns={'pagamento': 'payment_method'})

    return df

def _week_monday_friday_for(date_obj=None):
    """Retorna (monday_dt, friday_dt) datetimes (com extremos) para a semana que contém date_obj."""
    if date_obj is None:
        date_obj = datetime.now()
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.fromisoformat(date_obj)
        except Exception:
            date_obj = datetime.now()
    d = date_obj.date() if isinstance(date_obj, datetime) else date_obj
    monday = d - timedelta(days=d.weekday())
    friday = monday + timedelta(days=4)
    monday_dt = datetime.combine(monday, datetime.min.time())
    friday_dt = datetime.combine(friday, datetime.max.time())
    return monday_dt, friday_dt

def _format_excel_file(path):
    """Formata o xlsx para ficar mais apresentável (negrito header, ajustar colunas, freeze pane, moeda)."""
    try:
        wb = load_workbook(path)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill(start_color="FF2E7D32", end_color="FF2E7D32", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                try:
                    v = cell.value
                    ln = len(str(v)) if v is not None else 0
                except Exception:
                    ln = 0
                if ln > max_len: max_len = ln
            width = min(max(10, max_len + 2), 60)
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

        col_names = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        money_fmt = '"R$"#,##0.00'
        for name in ('price','valor_estoque','total_price'):
            if name in col_names:
                col_idx = col_names[name]
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = money_fmt
                    except Exception:
                        pass

        wb.save(path)
        return True
    except Exception as e:
        print(f"[WARN] Falha ao formatar Excel ({path}): {e}")
        return False

def generate_weekly_report(for_date=None):
    """
    Gera relatório semanal (Segunda a Sexta) com colunas: Segunda..Sexta, total_semana, estoque_atual, valor_estoque.
    Arquivo gerado em RELATORIO_DIR com nome baseado na **segunda-feira** da semana:
      Relatorios/relatorio_semanal_YYYY-MM-DD.xlsx  (YYYY-MM-DD = data da segunda-feira)
    Retorna caminho do arquivo gerado.
    """
    # aceita for_date como string ou datetime
    if isinstance(for_date, str):
        try:
            for_date = datetime.fromisoformat(for_date)
        except Exception:
            for_date = datetime.now()
    if for_date is None:
        for_date = datetime.now()

    monday_dt, friday_dt = _week_monday_friday_for(for_date)
    monday_date = monday_dt.date().isoformat()   # USAR a data da segunda-feira no nome do arquivo
    out_file = os.path.join(RELATORIO_DIR, f"relatorio_semanal_{monday_date}.xlsx")

    # lê todas as saidas normalizadas
    df_saidas = _read_saidas()

    # monta DataFrame base com todos os produtos (usa product_name como principal)
    products = load_products()
    prod_df = pd.DataFrame(products).rename(columns={
        'name': 'product_name',
        'id': 'product_id',
        'quantity': 'estoque_atual',
        'price': 'price'
    })
    # garante colunas mínimas
    for col in ['product_id', 'product_name', 'estoque_atual', 'price']:
        if col not in prod_df.columns:
            prod_df[col] = "" if col == 'product_name' else 0

    # se não houver saídas registradas -> gera planilha com zeros
    if df_saidas.empty:
        prod_df['Segunda'] = 0; prod_df['Terca'] = 0; prod_df['Quarta'] = 0; prod_df['Quinta'] = 0; prod_df['Sexta'] = 0
        prod_df['total_semana'] = 0
        prod_df['valor_estoque'] = prod_df['estoque_atual'].fillna(0).astype(float) * prod_df['price'].fillna(0).astype(float)
        out_df = prod_df[['product_name','product_id','Segunda','Terca','Quarta','Quinta','Sexta','total_semana','estoque_atual','price','valor_estoque']]
        try:
            out_df.to_excel(out_file, index=False, engine="openpyxl")
            _format_excel_file(out_file)
            print(f"[REPORT] Relatório semanal (vazio) gravado em: {out_file}")
            return out_file
        except Exception as e:
            csv_out = out_file.replace('.xlsx', '.csv')
            out_df.to_csv(csv_out, index=False, encoding='utf-8-sig')
            print(f"[WARN] Falha ao gravar XLSX ({e}). Gravado CSV em: {csv_out}")
            return csv_out

    # filtra período Monday..Friday
    df_week = df_saidas[(df_saidas['date_dt'] >= monday_dt) & (df_saidas['date_dt'] <= friday_dt)].copy()
    if df_week.empty:
        prod_df['Segunda'] = 0; prod_df['Terca'] = 0; prod_df['Quarta'] = 0; prod_df['Quinta'] = 0; prod_df['Sexta'] = 0
        prod_df['total_semana'] = 0
        prod_df['valor_estoque'] = prod_df['estoque_atual'].fillna(0).astype(float) * prod_df['price'].fillna(0).astype(float)
        out_df = prod_df[['product_name','product_id','Segunda','Terca','Quarta','Quinta','Sexta','total_semana','estoque_atual','price','valor_estoque']]
        try:
            out_df.to_excel(out_file, index=False, engine="openpyxl")
            _format_excel_file(out_file)
            print(f"[REPORT] Relatório semanal (semana sem vendas) gravado em: {out_file}")
            return out_file
        except Exception as e:
            csv_out = out_file.replace('.xlsx', '.csv')
            out_df.to_csv(csv_out, index=False, encoding='utf-8-sig')
            print(f"[WARN] Falha ao gravar XLSX ({e}). Gravado CSV em: {csv_out}")
            return csv_out

    # garantir que existe coluna product_name: se ausente, mapear via product_id
    if 'product_name' not in df_week.columns or df_week['product_name'].isnull().all():
        # tenta mapear product_id -> name
        prod_map = {str(p.get('id')): p.get('name') for p in products}
        df_week['product_name'] = df_week['product_id'].astype(str).map(prod_map).fillna(df_week.get('product_name',''))

    # pivot por product_name e weekday
    df_week['weekday'] = df_week['date_dt'].dt.weekday  # 0=Mon
    df_week = df_week[df_week['weekday'].isin([0,1,2,3,4])]
    df_week['qty_num'] = pd.to_numeric(df_week.get('qty_num', 0), errors='coerce').fillna(0).astype(int)

    pivot = pd.pivot_table(df_week, values='qty_num', index=['product_name'], columns='weekday', aggfunc='sum', fill_value=0)
    weekday_map = {0:'Segunda',1:'Terca',2:'Quarta',3:'Quinta',4:'Sexta'}
    pivot = pivot.rename(columns=weekday_map)

    # garantir colunas de dias
    for day in ['Segunda','Terca','Quarta','Quinta','Sexta']:
        if day not in pivot.columns:
            pivot[day] = 0
    pivot = pivot.reset_index()

    # merge pivot (por product_name) com lista completa de produtos (prod_df)
    merged = prod_df.merge(pivot, how='left', on='product_name')

    # preencher NaN dos dias com 0 e garantir tipos inteiros
    for day in ['Segunda','Terca','Quarta','Quinta','Sexta']:
        merged[day] = pd.to_numeric(merged.get(day, 0)).fillna(0).astype(int)

    merged['total_semana'] = merged[['Segunda','Terca','Quarta','Quinta','Sexta']].sum(axis=1)
    merged['valor_estoque'] = merged['estoque_atual'].fillna(0).astype(float) * merged['price'].fillna(0).astype(float)

    # define colunas finais com product_name primeiro para legibilidade
    out_cols = ['product_name','product_id','Segunda','Terca','Quarta','Quinta','Sexta','total_semana','estoque_atual','price','valor_estoque']
    out_df = merged[[c for c in out_cols if c in merged.columns]]

    # salva arquivo
    try:
        out_df.to_excel(out_file, index=False, engine="openpyxl")
        _format_excel_file(out_file)
        print(f"[REPORT] Relatório semanal gravado em: {out_file}")
        return out_file
    except Exception as e:
        csv_out = out_file.replace('.xlsx', '.csv')
        out_df.to_csv(csv_out, index=False, encoding='utf-8-sig')
        print(f"[WARN] Falha ao gravar XLSX ({e}). Gravado CSV em: {csv_out}")
        return csv_out

@app.route('/api/generate_weekly_report', methods=['GET','POST'])
def api_generate_weekly_report():
    """
    Gera (ou retorna existente) relatório semanal para a semana de 'date' (opcional).
    GET: aceita ?date=YYYY-MM-DD
    POST: aceita JSON { "date": "YYYY-MM-DD" }
    Retorna JSON {"ok": True, "file": "Relatorios/relatorio_semanal_YYYY-MM-DD.xlsx"}
    """
    d = None
    if request.method == 'GET':
        d = request.args.get('date')
    else:
        try:
            payload = request.get_json(force=True)
            d = payload.get('date') if isinstance(payload, dict) else None
        except Exception:
            d = None

    try:
        path = generate_weekly_report(for_date=d)
        fname = os.path.basename(path)
        return jsonify({"ok": True, "file": os.path.join(RELATORIO_DIR, fname)}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/reports/weekly', methods=['GET'])
def api_reports_weekly():
    """
    Retorna os dados agregados da semana (JSON) usando product_name como chave principal.
    Query param opcional: week_start=YYYY-MM-DD (data da segunda-feira)
    """
    week_start_str = request.args.get("week_start")
    if week_start_str:
        try:
            week_start = datetime.fromisoformat(week_start_str).date()
        except Exception:
            return jsonify({"error":"week_start inválido, use YYYY-MM-DD"}), 400
    else:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())  # Monday

    # lê SAIDAS_FILE
    df = _read_saidas()
    if df.empty:
        return jsonify({"week_start": str(week_start), "items": [], "note":"nenhuma saída registrada ainda"}), 200

    start_dt = datetime.combine(week_start, datetime.min.time())
    end_dt = datetime.combine(week_start + timedelta(days=4), datetime.max.time())
    week_df = df[(df['date_dt'] >= start_dt) & (df['date_dt'] <= end_dt)].copy()
    if week_df.empty:
        return jsonify({"week_start": str(week_start), "items": [], "note":"nenhuma saída nesta semana"}), 200

    # garantir product_name
    products = load_products()
    prod_map_by_id = {str(p.get('id')): p.get('name') for p in products}
    if 'product_name' not in week_df.columns or week_df['product_name'].isnull().all():
        week_df['product_name'] = week_df['product_id'].astype(str).map(prod_map_by_id).fillna('')

    week_df['weekday'] = week_df['date_dt'].dt.weekday
    week_df['qty_num'] = pd.to_numeric(week_df.get('qty_num', 0), errors='coerce').fillna(0).astype(int)

    pivot = week_df.groupby(['product_name','weekday'])['qty_num'].sum().reset_index()

    # agrega por produto
    products_out = {}
    for _, row in pivot.iterrows():
        pname = row['product_name'] or 'Desconhecido'
        wd = int(row['weekday'])
        qty = int(row['qty_num'])
        if pname not in products_out:
            products_out[pname] = {"product_name": pname, "by_day": {i:0 for i in range(5)}, "total":0}
        if 0 <= wd <= 4:
            products_out[pname]["by_day"][wd] += qty
            products_out[pname]["total"] += qty

    # enriquece com estoque/avg_cost
    prod_map_full = {p.get('name'): p for p in products}
    out = []
    for pname, info in products_out.items():
        p = prod_map_full.get(pname)
        stock = p.get("quantity") if p else None
        avg_cost = p.get("avg_cost") if p else None
        out.append({
            "product_name": pname,
            "mon": int(info["by_day"].get(0,0)),
            "tue": int(info["by_day"].get(1,0)),
            "wed": int(info["by_day"].get(2,0)),
            "thu": int(info["by_day"].get(3,0)),
            "fri": int(info["by_day"].get(4,0)),
            "weekly_total": int(info["total"]),
            "stock": stock,
            "avg_cost": avg_cost
        })

    # grava resumo opcional no RELATORIO_DIR com nome da monday
    try:
        report_df = pd.DataFrame(out)
        report_path = os.path.join(RELATORIO_DIR, f"weekly_report_{week_start.isoformat()}.xlsx")
        report_df.to_excel(report_path, index=False, engine="openpyxl")
        _format_excel_file(report_path)
    except Exception:
        report_path = None

    return jsonify({"week_start": str(week_start), "items": out, "report_file": report_path}), 200

# Atualiza automaticamente relatório da semana atual caso SAIDAS_FILE seja alterado
def append_rows_to_excel(path, rows, columns_order=None):
    """Anexa linhas (lista de dicts) a um arquivo excel; cria se não existir."""
    with file_lock:
        if os.path.exists(path):
            try:
                existing = pd.read_excel(path, engine="openpyxl")
            except Exception:
                # fallback try csv
                try:
                    existing = pd.read_csv(path.replace('.xlsx', '.csv'), dtype=str)
                except Exception:
                    existing = pd.DataFrame()
        else:
            existing = pd.DataFrame()

        new_df = pd.DataFrame(rows)
        if existing is None or existing.empty:
            out = new_df
        else:
            out = pd.concat([existing, new_df], ignore_index=True, sort=False)

        # reorder columns if requested
        if columns_order:
            for c in columns_order:
                if c not in out.columns:
                    out[c] = ""
            out = out[columns_order]

        try:
            out.to_excel(path, index=False, engine="openpyxl")
        except Exception as e:
            # fallback to csv
            out.to_csv(path.replace('.xlsx', '.csv'), index=False, encoding='utf-8-sig')
            print(f"[WARN] Falha ao gravar {path} ({e}), gravado CSV fallback.")
        finally:
            # se alteramos o arquivo de SAIDAS, atualiza o relatório da semana corrente
            try:
                if os.path.abspath(path) == os.path.abspath(SAIDAS_FILE):
                    # atualiza o relatório da semana corrente
                    try:
                        generate_weekly_report(for_date=datetime.now())
                    except Exception as e:
                        print("[WARN] Falha ao atualizar relatório weekly após append em SAIDAS:", e)
            except Exception:
                pass

# gera/atualiza relatório da semana atual no startup
try:
    generate_weekly_report(for_date=datetime.now())
except Exception as e:
    print("[WARN] Não foi possível gerar relatório weekly no startup:", e)

# Função utilitária para inserir novo produto em relatórios existente
def _ensure_product_in_report_excel(path, product):
    """
    Garante que o produto (dict com 'id','name','price','quantity') exista na planilha xlsx.
    Se não existir, adiciona uma linha ao final com zeros nos dias.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        # encontrar colunas básicas (product_id, product_name...). Se não existir, cria colunas mínimas.
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # map name->col idx
        col_index = {h: i+1 for i, h in enumerate(headers) if h}
        # decide se existe por nome
        exists = False
        pname = str(product.get('name', '')).strip()
        pid = str(product.get('id', ''))
        # busca por product_name na coluna
        if 'product_name' in col_index:
            col = col_index['product_name']
            for row in ws.iter_rows(min_row=2, values_only=True):
                v = row[col-1]
                if v and str(v).strip().lower() == pname.lower():
                    exists = True
                    break
        # se não existir, inserimos linha com zeros
        if not exists:
            # garante colunas esperadas existam
            expected = ['product_id','product_name','Segunda','Terca','Quarta','Quinta','Sexta','total_semana','estoque_atual','price','valor_estoque']
            # se faltar cabeçalho, adiciona colunas novas na planilha (após última coluna)
            last_col = ws.max_column
            hdrs = [str(c.value) if c.value is not None else "" for c in ws[1]]
            for ex in expected:
                if ex not in hdrs:
                    last_col += 1
                    ws.cell(row=1, column=last_col, value=ex)
            # recalcula hdrs mapping
            hdrs = [str(c.value) if c.value is not None else "" for c in ws[1]]
            hdr_map = {h: i+1 for i, h in enumerate(hdrs) if h}
            # nova linha de zeros
            new_row = {}
            new_row[hdr_map.get('product_id')] = pid
            new_row[hdr_map.get('product_name')] = pname
            for day in ['Segunda','Terca','Quarta','Quinta','Sexta']:
                new_row[hdr_map.get(day)] = 0
            new_row[hdr_map.get('total_semana')] = 0
            new_row[hdr_map.get('estoque_atual')] = int(product.get('quantity', 0) or 0)
            new_row[hdr_map.get('price')] = float(product.get('price', 0) or 0)
            new_row[hdr_map.get('valor_estoque')] = new_row[hdr_map.get('estoque_atual')] * new_row[hdr_map.get('price')]
            # append row at the end
            row_idx = ws.max_row + 1
            for col_idx, val in new_row.items():
                if col_idx:
                    ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(path)
        return True
    except Exception as e:
        print(f"[WARN] falha _ensure_product_in_report_excel {path}: {e}")
        return False

def add_product_to_all_reports(product):
    """Procura todos os arquivos relatorio_semanal_*.xlsx em RELATORIO_DIR e garante o produto em cada um."""
    try:
        pattern = os.path.join(RELATORIO_DIR, "relatorio_semanal_*.xlsx")
        files = glob.glob(pattern)
        for f in files:
            with file_lock:
                _ensure_product_in_report_excel(f, product)
        return True
    except Exception as e:
        print("[WARN] add_product_to_all_reports falhou:", e)
        return False

@app.route("/api/products", methods=["POST"])
def api_products_create():
    data = request.get_json(force=True)
    if not data or not data.get("name"):
        return jsonify({"error":"nome obrigatório"}), 400
    prods = load_products()
    max_id = max([p.get("id",0) for p in prods]) if prods else 0
    new_id = max_id + 1
    name = data.get("name")
    price = float(data.get("price", 0) or 0)
    qty = int(data.get("quantity", 0) or 0)
    category = data.get("category", "Outros")
    avg_cost = float(data.get("avg_cost", price or 0))
    new = {"id": new_id, "name": name, "price": price, "quantity": qty, "category": category, "avg_cost": avg_cost}
    prods.append(new)
    save_products(prods)
    # insere em relatórios já existentes
    try:
        add_product_to_all_reports(new)
    except Exception as e:
        print("[WARN] falha ao inserir produto novo nos relatórios existentes:", e)
    # emit product created
    try:
        socketio.emit("product_updated", {"action":"create","product":new})
    except Exception:
        pass
    return jsonify(new), 201


def encontrar_produto_nas_planilhas(nome_produto, fuzzy_threshold=0.68):
    nome_norm = normalize_text(nome_produto)
    if MANUAL_MAPPINGS:
        if nome_produto in MANUAL_MAPPINGS:
            m = MANUAL_MAPPINGS[nome_produto]
            print(f"[MAP] encontrado em manual_mappings para '{nome_produto}': {m}")
            return m

    best_candidate = None
    best_score = 0.0

    for path in EXCEL_DIR_FILES:
        if not os.path.exists(path):
            continue
        try:
            xls = pd.read_excel(path, sheet_name=None, dtype=str, engine="openpyxl")
        except Exception as e:
            print(f"[WARN] não conseguiu abrir {path}: {e}")
            continue

        for sheet_name, df in xls.items():
            for pos_row, (_, row) in enumerate(df.iterrows()):
                row_vals = row.values.astype(str)
                for pos_col, cell in enumerate(row_vals):
                    cell_text = "" if cell is None else str(cell)
                    cell_norm = normalize_text(cell_text)
                    if not cell_norm:
                        continue

                    if cell_norm == nome_norm:
                        print(f"[FOUND exact] '{nome_produto}' -> {path} | {sheet_name} | row {pos_row} col {pos_col}")
                        return {"file": path, "sheet": sheet_name, "row_idx": pos_row, "col_idx": pos_col, "matched_text": cell_text, "score": 1.0}

                    if nome_norm in cell_norm or cell_norm in nome_norm:
                        print(f"[FOUND substr] '{nome_produto}' ~ '{cell_text}' -> {path} | {sheet_name} | row {pos_row} col {pos_col}")
                        return {"file": path, "sheet": sheet_name, "row_idx": pos_row, "col_idx": pos_col, "matched_text": cell_text, "score": 0.9}

                    score = difflib.SequenceMatcher(None, nome_norm, cell_norm).ratio()
                    if score > best_score:
                        best_score = score
                        best_candidate = {"file": path, "sheet": sheet_name, "row_idx": pos_row, "col_idx": pos_col, "matched_text": cell_text, "score": score}

    if best_candidate and best_score >= fuzzy_threshold:
        print(f"[FOUND fuzzy {best_score:.2f}] '{nome_produto}' ~ '{best_candidate['matched_text']}' -> {best_candidate['file']} | {best_candidate['sheet']} | row {best_candidate['row_idx']} col {best_candidate['col_idx']}")
        return best_candidate

    print(f"[NOT MAPPED] produto não mapeado em planilhas: {nome_produto} (best_score={best_score:.2f})")
    return None

def decrementar_na_planilha(mapping, amount=1):
    if not mapping:
        return False, "mapping vazio"
    path = mapping["file"]
    sheet = mapping["sheet"]
    row = int(mapping["row_idx"])
    col = int(mapping["col_idx"])
    try:
        xls = pd.read_excel(path, sheet_name=None, engine="openpyxl")
        if sheet not in xls:
            return False, f"sheet '{sheet}' não encontrada em {path}"
        df = xls[sheet]
        val = df.iat[row, col]
        raw = "" if val is None else str(val).strip()
        if raw == "":
            return False, f"célula vazia em {path} [{sheet}] ({row},{col})"
        vnorm = re.sub(r'[^\d,.\-]', '', raw)
        vnorm = vnorm.replace(',', '.')
        try:
            num = float(vnorm)
        except Exception:
            return False, f"valor não numérico ({raw}) em {path} [{sheet}] ({row},{col})"
        novo = num - amount
        if novo < 0:
            novo = 0
        if float(novo).is_integer():
            df.iat[row, col] = int(novo)
        else:
            df.iat[row, col] = novo

        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            for sname, sheetdf in xls.items():
                sheetdf.to_excel(writer, sheet_name=sname, index=False)
        return True, f"novo={df.iat[row,col]} em {path} [{sheet}] ({row},{col})"
    except Exception as e:
        return False, str(e)

# ---------------- Rotas / API de produtos ----------------
@app.route("/api/products", methods=["GET"])
def api_products():
    prods = load_products()
    return jsonify(prods)

@app.route("/api/products", methods=["GET", "POST"])
def api_products_handler():
    """
    GET  -> retorna lista de produtos (JSON)
    POST -> cria novo produto. Body JSON: { "name": "...", "price": 1.23, "quantity": 10, "category": "X", "avg_cost": 1.0 }
    """
    if request.method == "GET":
        prods = load_products()
        return jsonify(prods)

    # POST -> create
    data = request.get_json(force=True)
    if not data or not data.get("name"):
        return jsonify({"error": "nome obrigatório"}), 400

    prods = load_products()
    try:
        max_id = max([int(p.get("id", 0)) for p in prods]) if prods else 0
    except Exception:
        max_id = 0
    new_id = max_id + 1

    name = data.get("name")
    price = float(data.get("price", 0) or 0)
    qty = int(data.get("quantity", 0) or 0)
    category = data.get("category", "Outros")
    avg_cost = float(data.get("avg_cost", price or 0))

    new = {
        "id": new_id,
        "name": name,
        "price": price,
        "quantity": qty,
        "category": category,
        "avg_cost": avg_cost
    }

    prods.append(new)
    save_products(prods)

    # notifica clientes conectados
    try:
        socketio.emit("product_updated", {"action": "create", "product": new})
    except Exception:
        pass

    return jsonify(new), 201


@app.route("/api/products/<int:pid>", methods=["PUT"])
def api_products_update(pid):
    data = request.get_json(force=True)
    prods = load_products()
    for p in prods:
        if int(p.get("id")) == int(pid):
            # update allowed fields
            for k in ["name","price","quantity","category","avg_cost"]:
                if k in data:
                    if k in ["price","avg_cost"]:
                        p[k] = float(data[k] or 0)
                    elif k == "quantity":
                        p[k] = int(data[k] or 0)
                    else:
                        p[k] = data[k]
            save_products(prods)
            try:
                socketio.emit("product_updated", {"action":"update","product":p})
            except Exception:
                pass
            return jsonify(p)
    return jsonify({"error":"produto não encontrado"}), 404

@app.route("/api/products/<int:pid>", methods=["DELETE"])
def api_products_delete(pid):
    prods = load_products()
    new_prods = [p for p in prods if int(p.get("id")) != int(pid)]
    if len(new_prods) == len(prods):
        return jsonify({"error":"produto não encontrado"}), 404
    save_products(new_prods)
    try:
        socketio.emit("product_updated", {"action":"delete","id":pid})
    except Exception:
        pass
    return jsonify({"ok": True})

@app.route("/api/products/<int:pid>/stock", methods=["POST"])
def api_products_stock(pid):
    """
    Ajusta estoque de um produto.
    Body:
      { "delta": +10 (positivo para entrada, negativo pra saida),
        "type": "entrada"|"saida",
        "supplier": "Fornecedor X",
        "unit_price": 3.2,
    """
    data = request.get_json(force=True)
    if not data or "delta" not in data:
        return jsonify({"error":"delta obrigatório"}), 400
    delta = int(data.get("delta", 0))
    typ = data.get("type", "entrada")
    supplier = data.get("supplier", "")
    unit_price = float(data.get("unit_price", 0) or 0)

    prods = load_products()
    target = None
    for p in prods:
        if int(p.get("id")) == int(pid):
            target = p
            break
    if not target:
        return jsonify({"error":"produto não encontrado"}), 404

    prev_qty = int(target.get("quantity", 0) or 0)
    
    if typ == "entrada":
        # recalcular custo médio
        incoming_qty = max(0, delta)
        if incoming_qty <= 0:
            return jsonify({"error":"delta deve ser positivo para entradas"}), 400
        total_incoming = incoming_qty * unit_price
        new_qty = prev_qty + incoming_qty
        target["quantity"] = new_qty
        # grava entrada na planilha entradas.xlsx
        row = {
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "product_id": target["id"],
            "product_name": target["name"],
            "supplier": supplier,
            "qty": incoming_qty,
            "unit_price": unit_price,
            "total_price": round(incoming_qty * unit_price, 4),
            "prev_qty": prev_qty,
            "new_qty": new_qty,
        }
        append_rows_to_excel(ENTRADAS_FILE, [row], columns_order=[
            "date","product_id","product_name","supplier","qty","unit_price","total_price",
            "prev_qty","new_qty"
        ])
    else:
        # saída: delta expected positive meaning how many taken out; or delta negative accepted too
        out_qty = abs(delta)
        if out_qty <= 0:
            return jsonify({"error":"delta inválido para saida"}), 400
        new_qty = max(0, prev_qty - out_qty)
        target["quantity"] = new_qty
        # tenta extrair método de pagamento de possíveis campos no payload
        payment_method = data.get("payment_method") or data.get("pagamento") or data.get("payment") or ""
        # grava saida na planilha saidas.xlsx
        row = {
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "product_id": target["id"],
            "product_name": target["name"],
            "qty": out_qty,
            "unit_price": unit_price,
            "total_price": round(out_qty * unit_price, 4),
            "prev_qty": prev_qty,
            "new_qty": new_qty,
            "payment_method": payment_method,
        }
        append_rows_to_excel(SAIDAS_FILE, [row], columns_order=[
            "date","product_id","product_name","qty","unit_price","total_price","prev_qty","new_qty","payment_method"
        ])


    save_products(prods)
    try:
        socketio.emit("product_updated", {"action":"stock","product":target})
    except Exception:
        pass
    return jsonify({"ok": True, "product": target})

# ---------------- Rotas existentes para pedidos / pedidos.xlsx ----------------
@app.route("/api/pedidos")
def api_pedidos():
    garantir_arquivo_pedidos()
    try:
        df = pd.read_excel(PEDIDOS_FILE, engine="openpyxl")
    except Exception as e:
        csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
        print(f"[WARN] Falha ao ler {PEDIDOS_FILE}: {e}. Tentando {csv_fallback}")
        if os.path.exists(csv_fallback):
            df = pd.read_csv(csv_fallback, dtype=str)
        else:
            df = pd.DataFrame(columns=["id","nome_cliente","itens","total","pagamento","status","hora_pedido","hora_conclusao"])
    return df.to_json(orient="records", force_ascii=False)

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

@app.route('/api/orders/<order_id>', methods=['DELETE'])
def api_delete_order(order_id):
    """
    Remove pedido do arquivo PEDIDOS_FILE (xlsx/csv).
    order_id pode ser numérico ou string; fazemos coercion/strip.
    """
    garantir_arquivo_pedidos()
    # normalizar id: tenta int, senão string
    try:
        order_id_norm = int(str(order_id))
    except Exception:
        order_id_norm = str(order_id).strip()

    with file_lock:
        # tenta ler xlsx, senão csv
        try:
            df = pd.read_excel(PEDIDOS_FILE, engine="openpyxl")
            read_source = "xlsx"
        except Exception as e:
            csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
            if os.path.exists(csv_fallback):
                df = pd.read_csv(csv_fallback, dtype=str)
                read_source = "csv"
            else:
                return jsonify({"error": "Arquivo de pedidos indisponível"}), 500

        if 'id' not in df.columns:
            return jsonify({"error": "Arquivo de pedidos sem coluna 'id'"}), 500

        # normalizar coluna id para comparar
        try:
            df_id_numeric = pd.to_numeric(df['id'], errors='coerce')
            if isinstance(order_id_norm, int):
                mask = df_id_numeric == order_id_norm
            else:
                mask = df['id'].astype(str).str.strip() == str(order_id_norm).strip()
        except Exception:
            mask = df['id'].astype(str).str.strip() == str(order_id_norm).strip()

        if not getattr(mask, 'any', lambda: False)():
            return jsonify({"error": "Pedido não encontrado"}), 404

        # remove as linhas que batem
        remaining = df.loc[~mask].copy()

        # grava de volta: tenta xlsx, se falhar grava csv fallback
        try:
            remaining.to_excel(PEDIDOS_FILE, index=False, engine="openpyxl")
        except Exception as e_write:
            csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
            remaining.to_csv(csv_fallback, index=False, encoding='utf-8-sig')

    # emitir para painéis que pedido foi removido (broadcast)
    try:
        _safe_emit("pedido_removido", {"id": order_id_norm})
    except Exception:
        pass

    return jsonify({"ok": True, "id": order_id_norm}), 200

# ---------------- Socket.IO ----------------
# --- SOCKET helpers / handlers corrigidos ---

@socketio.on("connect")
def on_connect():
    sid = request.sid if hasattr(request, 'sid') else None
    print("cliente conectou", sid)
    emit("server_message", {"msg": "connected"})


def _safe_emit(event, payload):
    """Emite para todos; tenta broadcast=True se disponível, senão emite sem."""
    try:
        # algumas versões do Flask-SocketIO aceitam broadcast kw, outras não
        socketio.emit(event, payload, broadcast=True)
        print(f"[EMIT] emitido {event} (broadcast) id=", payload.get("id"))
    except TypeError:
        try:
            socketio.emit(event, payload)
            print(f"[EMIT] emitido {event} (no-broadcast) id=", payload.get("id"))
        except Exception as ex:
            print(f"[ERRO] falha ao emitir socket {event}: {ex}")
    except Exception as ex:
        print(f"[ERRO] falha geral ao emitir socket {event}: {ex}")


@socketio.on("novo_pedido")
def on_novo_pedido(data):
    """
    Recebe pedido vindo por socket e delega a process_order_payload,
    que já faz salvar, decrementar estoque e emitir 'pedido_recebido'.
    Emite 'pedido_confirmado' apenas para o remetente.
    """
    try:
        pedido = process_order_payload(data)
        # confirma apenas para o cliente que enviou
        emit("pedido_confirmado", {"id": pedido.get("id"), "hora": pedido.get("hora_pedido")})
        print(f"[NOVO_PEDIDO] recebido id={pedido.get('id')}")
    except Exception as e:
        print("[ERRO] on_novo_pedido:", e)
        try:
            emit("pedido_error", {"error": str(e)})
        except Exception:
            pass


@socketio.on("concluir_pedido")
def on_concluir_pedido(data):
    """
    Marca pedido como concluído no arquivo PEDIDOS_FILE (.xlsx/.csv).
    Suporta id como int ou string. Emite pedido_concluido (broadcast) e concluir_ok/err.
    """
    raw_id = data.get("id")
    if raw_id is None:
        emit("concluir_err", {"id": None, "msg": "id ausente"})
        return

    # normaliza: tenta int; se falhar, mantem string
    try:
        pedido_id = int(str(raw_id))
    except Exception:
        pedido_id = str(raw_id).strip()

    garantir_arquivo_pedidos()

    with file_lock:
        try:
            df = pd.read_excel(PEDIDOS_FILE, engine="openpyxl")
        except Exception as e:
            csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
            print(f"[WARN] Falha ao ler {PEDIDOS_FILE} para concluir pedido: {e}. Tentando {csv_fallback}")
            if os.path.exists(csv_fallback):
                df = pd.read_csv(csv_fallback, dtype=str)
            else:
                emit("concluir_err", {"id": raw_id, "msg": "Arquivo de pedidos indisponível"})
                return

        if 'id' not in df.columns:
            emit("concluir_err", {"id": raw_id, "msg": "Arquivo de pedidos sem coluna 'id'."})
            return

        # tentar comparação numérica quando possível
        mask = None
        try:
            df_id_numeric = pd.to_numeric(df['id'], errors='coerce')
            if isinstance(pedido_id, int):
                mask = df_id_numeric == pedido_id
            else:
                mask = df['id'].astype(str).str.strip() == str(pedido_id).strip()
        except Exception:
            mask = df['id'].astype(str).str.strip() == str(pedido_id).strip()

        if not getattr(mask, 'any', lambda: False)():
            emit("concluir_err", {"id": raw_id, "msg": "Pedido não encontrado"})
            return

        try:
            df.loc[mask, "status"] = "concluido"
            df.loc[mask, "hora_conclusao"] = datetime.now().isoformat(sep=' ', timespec='seconds')
            try:
                df.to_excel(PEDIDOS_FILE, index=False, engine="openpyxl")
            except Exception as e_write:
                csv_fallback = PEDIDOS_FILE.replace('.xlsx', '.csv')
                df.to_csv(csv_fallback, index=False, encoding='utf-8-sig')
                print(f"[WARN] Não foi possível gravar .xlsx ({e_write}). Gravado CSV fallback em: {csv_fallback}")
            # notifica painéis e solicita o OK para o solicitante
            _safe_emit("pedido_concluido", {"id": pedido_id})
            emit("concluir_ok", {"id": pedido_id})
            print(f"[CONCLUIR] pedido {pedido_id} marcado como concluído")
        except Exception as e:
            print("[ERRO] ao marcar pedido concluído:", e)
            emit("concluir_err", {"id": raw_id, "msg": str(e)})


def process_order_payload(data):
    """
    Cria o pedido (normalizando buyerInfo/name, pagamento, note),
    salva em planilha, decrementa estoque, registra saídas e emite pedido_recebido.
    Retorna o objeto pedido.
    """
    pedido_id = int(datetime.now().timestamp() * 1000)
    hora = datetime.now().isoformat(sep=' ', timespec='seconds')

    # extrair observações (aceita várias chaves)
    note = ''
    if isinstance(data, dict):
        note = (
            data.get('note')
            or data.get('notes')
            or data.get('observacoes')
            or data.get('observacao')
            or data.get('observações')
            or data.get('obs')
            or (data.get('buyerInfo') or {}).get('note')
            or ''
        )
    if note is None:
        note = ''
    else:
        note = str(note)

    pagamento = data.get("pagamento") or data.get("payment") or data.get("paymentMethod") or data.get("payment_method") or ""

    pedido = {
        "id": pedido_id,
        "nome_cliente": (data.get("buyerInfo", {}) or {}).get("name") or data.get("nome_cliente") or "",
        "itens": data.get("itens") or [],
        "total": data.get("total"),
        "pagamento": pagamento,
        "status": "novo",
        "hora_pedido": hora,
        "hora_conclusao": "",
        "note": note
    }

    # salvar pedido (função existente)
    salvar_pedido_excel(pedido)

    # decrementar estoque e registrar saidas
    products = load_products() or []
    saidas_rows = []
    payment_method = pagamento or data.get("payment_method") or ""

    for it in data.get("itens", []):
        # normalizar item (objeto ou string)
        if isinstance(it, str):
            try:
                it_obj = json.loads(it)
            except Exception:
                it_obj = {"name": it, "qtd": 1}
        else:
            it_obj = it

        name = it_obj.get("name") or it_obj.get("nome") or ""
        try:
            qtd = int(it_obj.get("qtd", it_obj.get("qty", it_obj.get("quantity", 1))) or 1)
        except Exception:
            qtd = 1

        matched = False
        for p in products:
            if (p.get("name") or "").strip().lower() == (name or "").strip().lower():
                prev_q = int(p.get("quantity", 0) or 0)
                p["quantity"] = max(0, prev_q - qtd)
                saidas_rows.append({
                    "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "product_id": p.get("id"),
                    "product_name": p.get("name"),
                    "qty": qtd,
                    "unit_price": p.get("price", 0),
                    "total_price": round(qtd * (p.get("price", 0) or 0), 4),
                    "prev_qty": prev_q,
                    "new_qty": p.get("quantity"),
                    "payment_method": payment_method,
                })
                matched = True

        mapping = encontrar_produto_nas_planilhas(name)
        if mapping:
            try:
                ok, msg = decrementar_na_planilha(mapping, amount=qtd)
                print("decrement planilha:", ok, msg)
            except Exception as e:
                print("erro decrement planilha:", e)
        else:
            if not matched:
                print("produto não mapeado em planilhas nem products.json:", name)

    if saidas_rows:
        append_rows_to_excel(SAIDAS_FILE, saidas_rows, columns_order=[
            "date", "product_id", "product_name", "qty", "unit_price", "total_price", "prev_qty", "new_qty", "payment_method"
        ])

    save_products(products)

    # emitir pedido_recebido (broadcast) com campo note incluso
    try:
        _safe_emit("pedido_recebido", pedido)
    except Exception as e:
        print("[WARN] falha ao emitir pedido_recebido:", e)

    return pedido


@app.route('/api/orders', methods=['POST'])
def api_orders():
    # usa silent para não levantar exceção quando JSON inválido
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "payload vazio ou JSON inválido"}), 400
    pedido = process_order_payload(data)
    # caso a chamada seja HTTP, também retornamos confirmação imediata
    return jsonify({"ok": True, "order": pedido}), 201

# ---------------- Entradas / Saidas endpoints (HTTP convenientes) ----------------
@app.route('/api/entradas', methods=['POST'])
def api_entradas():
    """
    Body example:
    {
      "product_id": 3,
      "qty": 10,
      "unit_price": 2.5,
      "supplier": "Fornecedor X",
      "note": "Compra julho"
    }
    """
    data = request.get_json(force=True)
    if not data:
        return jsonify({"error":"payload vazio"}), 400
    pid = data.get("product_id")
    qty = int(data.get("qty",0) or 0)
    unit_price = float(data.get("unit_price",0) or 0)
    supplier = data.get("supplier","")
    note = data.get("note","")
    if not pid or qty <= 0:
        return jsonify({"error":"product_id e qty (>0) necessários"}), 400
    # usa endpoint de stock com delta positivo
    resp = api_products_stock.__wrapped__(pid) if hasattr(api_products_stock, "__wrapped__") else None
    # chamar diretamente a lógica reutilizando a função de ajuste (evita duplicação)
    return api_products_stock(pid)

@app.route('/api/saidas', methods=['POST'])
def api_saidas():
    data = request.get_json(force=True)
    if not data:
        return jsonify({"error":"payload vazio"}), 400
    pid = data.get("product_id")
    qty = int(data.get("qty",0) or 0)
    note = data.get("note","")
    if not pid or qty <= 0:
        return jsonify({"error":"product_id e qty (>0) necessários"}), 400
    # chama stock com delta negativo
    payload = {"delta": -abs(qty), "type": "saida", "unit_price": data.get("unit_price", 0), "note": note, "source": data.get("source","manual")}
    with app.test_request_context(json=payload):
        return api_products_stock(pid)

# ---------------- Relatórios ----------------
def get_monday(d):
    # recebe date, retorna monday date
    return d - timedelta(days=d.weekday())

# rota adicional mais compacta (JSON por API): /api/reports/weekly_summary?week_start=YYYY-MM-DD
@app.route('/api/reports/weekly_summary', methods=['GET'])
def api_reports_weekly_summary():
    """
    Versão JSON do relatório semanal.
    Query param opcional: week_start=YYYY-MM-DD (segunda-feira)
    Retorna {"week_start": "...", "items": [...], "report_file": "..."}
    """
    week_start_str = request.args.get("week_start")
    if week_start_str:
        try:
            week_start = datetime.fromisoformat(week_start_str).date()
        except Exception:
            return jsonify({"error":"week_start inválido, use YYYY-MM-DD"}), 400
    else:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())

    # lê SAIDAS_FILE usando helper (mantém normalizações)
    df = _read_saidas()
    if df.empty:
        return jsonify({"week_start": str(week_start), "items": [], "note": "nenhuma saída registrada ainda"}), 200

    start_dt = datetime.combine(week_start, datetime.min.time())
    end_dt = datetime.combine(week_start + timedelta(days=4), datetime.max.time())
    week_df = df[(df['date_dt'] >= start_dt) & (df['date_dt'] <= end_dt)].copy()
    if week_df.empty:
        return jsonify({"week_start": str(week_start), "items": [], "note": "nenhuma saída nesta semana"}), 200

    # preparar e agregar por dia
    week_df['weekday'] = week_df['date_dt'].dt.weekday
    week_df['qty_num'] = pd.to_numeric(week_df.get('qty_num', 0), errors='coerce').fillna(0).astype(int)

    pivot = week_df.groupby(['product_id','product_name','weekday'])['qty_num'].sum().reset_index()

    products = {}
    for _, row in pivot.iterrows():
        pid_raw = row.get('product_id')
        pid = str(pid_raw) if not pd.isna(pid_raw) else str(row.get('product_name') or '')
        pname = row.get('product_name') or ''
        wd = int(row['weekday'])
        qty = int(row['qty_num'])
        if pid not in products:
            products[pid] = {"product_id": pid, "product_name": pname, "by_day": {i:0 for i in range(5)}, "total":0}
        if 0 <= wd <= 4:
            products[pid]["by_day"][wd] += qty
            products[pid]["total"] += qty

    # enriquece com estoque atual e avg_cost
    prods = load_products()
    prod_map = {str(p['id']): p for p in prods}

    out = []
    for pid, info in products.items():
        p = prod_map.get(pid)
        stock = p.get("quantity") if p else None
        avg_cost = p.get("avg_cost") if p else None
        out.append({
            "product_id": pid,
            "product_name": info["product_name"],
            "mon": int(info["by_day"].get(0,0)),
            "tue": int(info["by_day"].get(1,0)),
            "wed": int(info["by_day"].get(2,0)),
            "thu": int(info["by_day"].get(3,0)),
            "fri": int(info["by_day"].get(4,0)),
            "weekly_total": int(info["total"]),
            "stock": stock,
            "avg_cost": avg_cost
        })

    # grava planilha resumida semanal (opcional) em Relatorios/
    try:
        report_df = pd.DataFrame(out)
        os.makedirs(RELATORIO_DIR, exist_ok=True)
        report_path = os.path.join(RELATORIO_DIR, f"weekly_report_{week_start.isoformat()}.xlsx")
        report_df.to_excel(report_path, index=False, engine="openpyxl")
        _format_excel_file(report_path)
    except Exception:
        report_path = None

    return jsonify({"week_start": str(week_start), "items": out, "report_file": report_path}), 200

@app.route('/api/test_emit', methods=['GET'])
def api_test_emit():
    pedido_id = int(datetime.now().timestamp() * 1000)
    hora = datetime.now().isoformat(sep=' ', timespec='seconds')
    pedido = {
        "id": pedido_id,
        "nome_cliente": "TESTE SOCKET",
        "itens": '[{"name":"SNIKERS","qtd":1}]',
        "total": 1.0,
        "pagamento": "pix",
        "status": "novo",
        "hora_pedido": hora,
        "hora_conclusao": ""
    }
    try:
        socketio.emit('pedido_recebido', pedido)
        print(f"[EMIT TEST] emitido pedido_recebido id= {pedido_id}")
        return jsonify({"ok": True, "emitted": pedido}), 200
    except Exception as e:
        print("[EMIT TEST ERROR]", e)
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route('/_routes')
def _list_routes():
    lines = []
    for rule in sorted(app.url_map.iter_rules(), key=lambda r: r.rule):
        methods = ','.join(sorted(rule.methods))
        lines.append(f"{rule.rule}  -> endpoint={rule.endpoint}  methods={methods}")
    return "<pre>" + "\n".join(lines) + "</pre>"


@app.route('/')
def serve_root_index():
    # tenta index na pasta html
    p = os.path.join('html', 'index.html')
    if os.path.exists(p):
        return send_from_directory('html', 'index.html')
    # fallback: tenta html/index.html (caso path diferente)
    p2 = os.path.join(os.getcwd(), 'html', 'index.html')
    if os.path.exists(p2):
        return send_from_directory('html', 'index.html')
    abort(404)

# --- Serve páginas e assets estáticos (apenas UMA vez cada rota) ---
from flask import send_from_directory  # já deve estar importado no topo do app.py

# Serve qualquer outro arquivo que estiver na pasta html/ (ex: html/some.js)
@app.route('/html/<path:filename>', methods=['GET'])
def serve_html_asset(filename):
    return send_from_directory('html', filename)

# Serve scripts (pasta Script/)
@app.route('/script/<path:filename>', methods=['GET'])
def serve_script(filename):
    return send_from_directory('Script', filename)

# Serve estilos/imagens (pasta Style/)
@app.route('/style/<path:filename>', methods=['GET'])
def serve_style(filename):
    return send_from_directory('Style', filename)

# Serve imagens internas em Style/img/ (rota /img/...)
@app.route('/img/<path:filename>', methods=['GET'])
def serve_img(filename):
    return send_from_directory(os.path.join('Style', 'img'), filename)

# Serve arquivos de dados (Data/)
@app.route('/data/<path:filename>', methods=['GET'])
def serve_data(filename):
    return send_from_directory('Data', filename)

# Serve relatórios (Relatorios/)
@app.route('/relatorios/<path:filename>', methods=['GET'])
def serve_relatorio(filename):
    return send_from_directory('Relatorios', filename)




# ---------------- Main ----------------
if __name__ == "__main__":
    reparar_arquivo_pedidos()
    garantir_arquivo_pedidos()
    load_products()
    socketio.run(app, host="0.0.0.0", port=5000)
